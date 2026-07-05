import os
import sys
import glob
import json
import urllib.parse
import webbrowser
import fitz  # PyMuPDF
import cv2
import openpyxl

# Path Konfigurasi
BASE_DIR = r"C:\sertifikat_digital"
READER_DIR = os.path.join(BASE_DIR, "pembaca qr code")
EXCEL_PATH = os.path.join(BASE_DIR, "sertifikat.xlsx")
DATABASE_JS_PATH = os.path.join(BASE_DIR, "output", "database.js")
LOCAL_HTML_PATH = os.path.join(BASE_DIR, "output", "index.html")

def load_local_database():
    """Memuat database sertifikat dari database.js atau Excel untuk pencocokan"""
    database = {}
    
    # Prioritas 1: Ambil dari database.js (cepat & terformat)
    if os.path.exists(DATABASE_JS_PATH):
        try:
            with open(DATABASE_JS_PATH, "r", encoding="utf-8") as f:
                content = f.read()
                # Bersihkan pembungkus variabel JS 'const certificateDatabase = ...;'
                if "const certificateDatabase =" in content:
                    json_str = content.split("const certificateDatabase =")[1].strip()
                    if json_str.endswith(";"):
                        json_str = json_str[:-1].strip()
                    database = json.loads(json_str)
                    return database
        except Exception as e:
            print(f"Peringatan: Gagal memuat database.js ({e}), mencoba membaca Excel...")

    # Prioritas 2: Ambil langsung dari Excel sertifikat.xlsx
    if os.path.exists(EXCEL_PATH):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            idx_id = headers.index("Nomor Sertifikat")
            idx_penerima = headers.index("Penerima")
            idx_kegiatan = headers.index("Kegiatan")
            idx_sebagai = headers.index("Sebagai") if "Sebagai" in headers else -1
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= idx_id or not row[idx_id]:
                    continue
                id_cert = str(row[idx_id])
                database[id_cert] = {
                    "name": str(row[idx_penerima]) if row[idx_penerima] else "-",
                    "role": str(row[idx_sebagai]) if (idx_sebagai != -1 and row[idx_sebagai]) else "Peserta",
                    "activity": str(row[idx_kegiatan]) if row[idx_kegiatan] else "-",
                    "date": "Terdaftar di Excel"
                }
        except Exception as e:
            print(f"Error: Gagal memuat database dari Excel: {e}")
            
    return database

def decode_qr_from_image(image_path):
    """Membaca QR Code dari gambar menggunakan OpenCV"""
    img = cv2.imread(image_path)
    if img is None:
        return None
    
    # Inisialisasi detector QR OpenCV
    detector = cv2.QRCodeDetector()
    val, points, straight_qrcode = detector.detectAndDecode(img)
    
    if val:
        return val
    
    # Jika gagal, coba ubah ke grayscale dan tingkatkan kontras
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    val_gray, _, _ = detector.detectAndDecode(gray)
    if val_gray:
        return val_gray
        
    return None

def scan_pdf_for_qr(pdf_path):
    """Membuka PDF, merender halaman menjadi gambar, dan mendeteksi QR Code"""
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        print(f"Error: Gagal membuka file PDF: {e}")
        return None

    qr_data = None
    # Cek halaman satu per satu (biasanya halaman pertama)
    for page_idx in range(min(5, len(doc))):
        page = doc.load_page(page_idx)
        
        # Render halaman ke gambar (gunakan resolusi tinggi zoom=2.0 agar QR Code jelas)
        zoom = 2.0
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        
        # Simpan gambar temp
        temp_img_path = os.path.join(READER_DIR, f"temp_page_{page_idx}.png")
        pix.save(temp_img_path)
        
        # Scan gambar temp
        qr_data = decode_qr_from_image(temp_img_path)
        
        # Hapus file temp
        if os.path.exists(temp_img_path):
            os.remove(temp_img_path)
            
        if qr_data:
            break
            
    doc.close()
    return qr_data

def open_in_browser(url):
    """Membuka URL di browser Microsoft Edge secara eksplisit pada Windows"""
    print(f"\nMembuka di browser Microsoft Edge:")
    print(f"-> {url}")
    try:
        if sys.platform == "win32":
            # Jalankan Edge secara eksplisit dengan perintah shell
            os.system(f'start msedge "{url}"')
        else:
            webbrowser.open(url)
    except Exception as e:
        print(f"Gagal membuka browser otomatis: {e}")
        print("Silakan salin tautan di atas dan buka secara manual di browser Anda.")

def main():
    print("==================================================")
    print("      PEMBACA QR CODE SERTIFIKAT DIGITAL          ")
    print("==================================================")
    
    # Cari semua file PDF di folder 'pembaca qr code'
    pdf_files = glob.glob(os.path.join(READER_DIR, "*.pdf"))
    
    if not pdf_files:
        print("Tidak ada file PDF terdeteksi di folder:")
        print(f"-> {READER_DIR}")
        print("\nSilakan letakkan file PDF sertifikat di folder ini lalu jalankan skrip lagi.")
        print("==================================================")
        return

    # Muat database lokal untuk verifikasi metadata
    db = load_local_database()
    print(f"Memuat {len(db)} data sertifikat terdaftar...\n")

    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        print(f"Membaca file: {filename}...")
        
        # Scan PDF
        qr_url = scan_pdf_for_qr(pdf_path)
        
        if qr_url:
            print("-> [SUKSES] QR Code Terdeteksi!")
            print(f"-> URL QR Code: {qr_url}")
            
            # Coba ekstrak ID sertifikat dari URL (?id=...)
            cert_id = None
            try:
                parsed_url = urllib.parse.urlparse(qr_url)
                params = urllib.parse.parse_qs(parsed_url.query)
                if 'id' in params:
                    cert_id = params['id'][0]
            except Exception:
                pass
                
            if not cert_id:
                # Jika QR Code bukan format URL kita, cetak isi teks langsung
                cert_id = qr_url.strip()

            print("\n--- HASIL VERIFIKASI DIGITAL ---")
            if cert_id in db:
                data = db[cert_id]
                print(f"Nomor Sertifikat : {cert_id}")
                print(f"Nama Penerima    : {data['name']}")
                print(f"Sebagai          : {data['role']}")
                print(f"Nama Kegiatan    : {data['activity']}")
                print(f"Status           : [ VERIFIED / SAH & TERDAFTAR ]")
                
                # Buka di browser secara lokal agar user bisa melihat tampilannya langsung
                if os.path.exists(LOCAL_HTML_PATH):
                    local_url = f"file:///{LOCAL_HTML_PATH.replace('\\', '/')}?id={urllib.parse.quote(cert_id)}"
                    open_in_browser(local_url)
                else:
                    open_in_browser(qr_url)
            else:
                print(f"Nomor Sertifikat : {cert_id}")
                print("Status           : [ TIDAK TERDAFTAR DI DATABASE ]")
                print("Peringatan: QR Code valid tetapi ID sertifikat ini tidak ditemukan di database.")
        else:
            print("-> [GAGAL] Tidak menemukan QR Code pada file PDF ini.")
            
        print("==================================================\n")

if __name__ == "__main__":
    main()
